"""
VEA Connect - Crear Índice de Búsqueda con Azure AI Studio
Script para crear índice de búsqueda usando AIProjectClient (como el ejemplo)
"""

import os
import sys
from azure.ai.projects import AIProjectClient
from azure.ai.projects.models import ConnectionType
from azure.identity import DefaultAzureCredential
from azure.core.credentials import AzureKeyCredential
from azure.search.documents import SearchClient
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents.indexes.models import (
    SemanticSearch,
    SearchField,
    SimpleField,
    SearchableField,
    SearchFieldDataType,
    SemanticConfiguration,
    SemanticPrioritizedFields,
    SemanticField,
    VectorSearch,
    HnswAlgorithmConfiguration,
    VectorSearchAlgorithmKind,
    HnswParameters,
    VectorSearchAlgorithmMetric,
    ExhaustiveKnnAlgorithmConfiguration,
    ExhaustiveKnnParameters,
    VectorSearchProfile,
    SearchIndex,
)
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

def get_logger(name):
    """Logger simple para el script"""
    import logging
    logging.basicConfig(level=logging.INFO)
    return logging.getLogger(name)

logger = get_logger(__name__)

def create_index_definition(index_name: str, model: str = "text-embedding-ada-002") -> SearchIndex:
    """Crear definición del índice de búsqueda para VEA Connect"""
    
    # Dimensiones del vector según el modelo
    dimensions = 1536  # text-embedding-ada-002
    if model == "text-embedding-3-large":
        dimensions = 3072
    
    # Campos del índice adaptados para VEA Connect
    fields = [
        SimpleField(name="id", type=SearchFieldDataType.String, key=True),
        SearchableField(name="content", type=SearchFieldDataType.String),
        SearchableField(name="merged_content", type=SearchFieldDataType.String),
        SimpleField(name="filepath", type=SearchFieldDataType.String),
        SearchableField(name="title", type=SearchFieldDataType.String),
        SimpleField(name="url", type=SearchFieldDataType.String),
        SimpleField(name="metadata_storage_name", type=SearchFieldDataType.String),
        SimpleField(name="metadata_storage_path", type=SearchFieldDataType.String),
        SimpleField(name="metadata_storage_size", type=SearchFieldDataType.Int64),
        SimpleField(name="metadata_storage_last_modified", type=SearchFieldDataType.DateTimeOffset),
        SimpleField(name="language", type=SearchFieldDataType.String),
        SimpleField(name="document_type", type=SearchFieldDataType.String),
        SimpleField(name="category", type=SearchFieldDataType.String),
        SearchableField(name="keyphrases", type=SearchFieldDataType.Collection(SearchFieldDataType.String)),
        SearchableField(name="locations", type=SearchFieldDataType.Collection(SearchFieldDataType.String)),
        SearchableField(name="imageTags", type=SearchFieldDataType.Collection(SearchFieldDataType.String)),
        SearchableField(name="imageCaption", type=SearchFieldDataType.Collection(SearchFieldDataType.String)),
        SearchableField(name="text", type=SearchFieldDataType.Collection(SearchFieldDataType.String)),
        SearchableField(name="layoutText", type=SearchFieldDataType.Collection(SearchFieldDataType.String)),
        SearchField(
            name="contentVector",
            type=SearchFieldDataType.Collection(SearchFieldDataType.Single),
            searchable=True,
            vector_search_dimensions=dimensions,
            vector_search_profile_name="myHnswProfile",
        ),
    ]

    # Configuración semántica para VEA Connect
    semantic_config = SemanticConfiguration(
        name="default",
        prioritized_fields=SemanticPrioritizedFields(
            title_field=SemanticField(field_name="title"),
            keywords_fields=[SemanticField(field_name="keyphrases")],
            content_fields=[
                SemanticField(field_name="content"),
                SemanticField(field_name="merged_content")
            ],
        ),
    )

    # Configuración de búsqueda vectorial
    vector_search = VectorSearch(
        algorithms=[
            HnswAlgorithmConfiguration(
                name="myHnsw",
                kind=VectorSearchAlgorithmKind.HNSW,
                parameters=HnswParameters(
                    m=4,
                    ef_construction=1000,
                    ef_search=1000,
                    metric=VectorSearchAlgorithmMetric.COSINE,
                ),
            ),
            ExhaustiveKnnAlgorithmConfiguration(
                name="myExhaustiveKnn",
                kind=VectorSearchAlgorithmKind.EXHAUSTIVE_KNN,
                parameters=ExhaustiveKnnParameters(metric=VectorSearchAlgorithmMetric.COSINE),
            ),
        ],
        profiles=[
            VectorSearchProfile(
                name="myHnswProfile",
                algorithm_configuration_name="myHnsw",
            ),
            VectorSearchProfile(
                name="myExhaustiveKnnProfile",
                algorithm_configuration_name="myExhaustiveKnn",
            ),
        ],
    )

    # Configuración de búsqueda semántica
    semantic_search = SemanticSearch(configurations=[semantic_config])

    # Crear definición del índice
    return SearchIndex(
        name=index_name,
        fields=fields,
        semantic_search=semantic_search,
        vector_search=vector_search,
    )

def create_docs_from_blob_storage(embeddings_client, model: str) -> list[dict]:
    """Crear documentos desde Azure Blob Storage con embeddings vectoriales"""
    from azure.storage.blob import BlobServiceClient
    
    try:
        # Obtener conexión a Blob Storage
        storage_connection_string = os.environ.get("STORAGE_CONNECTION_STRING")
        container_name = os.environ.get("STORAGE_CONTAINER_NAME", "admin-documentos")
        
        if not storage_connection_string:
            logger.warning("No se encontró STORAGE_CONNECTION_STRING, creando índice vacío")
            return []
        
        blob_service_client = BlobServiceClient.from_connection_string(storage_connection_string)
        container_client = blob_service_client.get_container_client(container_name)
        
        items = []
        doc_id = 1
        
        # Listar todos los blobs en el contenedor
        blobs = container_client.list_blobs()
        
        for blob in blobs:
            try:
                # Obtener el blob
                blob_client = container_client.get_blob_client(blob.name)
                blob_data = blob_client.download_blob().readall()
                
                # Determinar tipo de contenido
                content_type = blob.properties.content_type or "application/octet-stream"
                
                # Extraer texto según el tipo de archivo
                content = extract_text_from_blob(blob_data, content_type, blob.name)
                
                if not content:
                    logger.warning(f"No se pudo extraer texto de {blob.name}")
                    continue
                
                # Generar embedding vectorial
                emb = embeddings_client.embed(input=content, model=model)
                
                # Determinar categoría basada en el nombre del archivo
                category = determine_category(blob.name)
                
                rec = {
                    "id": str(doc_id),
                    "content": content,
                    "merged_content": content,
                    "filepath": blob.name,
                    "title": get_title_from_filename(blob.name),
                    "url": f"/documentos/{blob.name}",
                    "metadata_storage_name": blob.name,
                    "metadata_storage_path": f"/{container_name}/{blob.name}",
                    "metadata_storage_size": blob.size,
                    "metadata_storage_last_modified": blob.last_modified.isoformat() if blob.last_modified else None,
                    "language": "es",
                    "document_type": get_document_type(content_type),
                    "category": category,
                    "contentVector": emb.data[0].embedding,
                }
                items.append(rec)
                doc_id += 1
                
            except Exception as e:
                logger.error(f"Error procesando blob {blob.name}: {str(e)}")
                continue
        
        logger.info(f"Procesados {len(items)} documentos desde Blob Storage")
        return items
        
    except Exception as e:
        logger.error(f"Error al conectar con Blob Storage: {str(e)}")
        return []

def extract_text_from_blob(blob_data: bytes, content_type: str, filename: str) -> str:
    """Extraer texto de diferentes tipos de archivos"""
    try:
        if content_type.startswith("text/"):
            return blob_data.decode("utf-8")
        
        elif content_type == "application/pdf":
            try:
                import PyPDF2
                import io
                pdf_reader = PyPDF2.PdfReader(io.BytesIO(blob_data))
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                return text.strip() if text.strip() else f"PDF: {filename}"
            except ImportError:
                return f"PDF: {filename} (PyPDF2 no instalado)"
            except Exception as e:
                logger.warning(f"Error extrayendo PDF {filename}: {str(e)}")
                return f"PDF: {filename}"
        
        elif content_type in ["application/msword", "application/vnd.openxmlformats-officedocument.wordprocessingml.document"]:
            try:
                from docx import Document
                import io
                doc = Document(io.BytesIO(blob_data))
                text = ""
                for paragraph in doc.paragraphs:
                    text += paragraph.text + "\n"
                return text.strip() if text.strip() else f"Word: {filename}"
            except ImportError:
                return f"Word: {filename} (python-docx no instalado)"
            except Exception as e:
                logger.warning(f"Error extrayendo Word {filename}: {str(e)}")
                return f"Word: {filename}"
        
        elif content_type == "application/vnd.ms-excel" or content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            try:
                import openpyxl
                import io
                workbook = openpyxl.load_workbook(io.BytesIO(blob_data))
                text = ""
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    text += f"Hoja: {sheet_name}\n"
                    for row in sheet.iter_rows(values_only=True):
                        if any(cell for cell in row if cell):
                            text += " | ".join(str(cell) for cell in row if cell) + "\n"
                return text.strip() if text.strip() else f"Excel: {filename}"
            except ImportError:
                return f"Excel: {filename} (openpyxl no instalado)"
            except Exception as e:
                logger.warning(f"Error extrayendo Excel {filename}: {str(e)}")
                return f"Excel: {filename}"
        
        elif content_type.startswith("image/"):
            try:
                return extract_text_from_image(blob_data, filename)
            except Exception as e:
                logger.warning(f"Error extrayendo texto de imagen {filename}: {str(e)}")
                return f"Imagen: {filename}"
        
        else:
            return f"Archivo: {filename}"
            
    except Exception as e:
        logger.error(f"Error extrayendo texto de {filename}: {str(e)}")
        return f"Error procesando: {filename}"

def determine_category(filename: str) -> str:
    """Determinar categoría basada en el nombre del archivo"""
    filename_lower = filename.lower()
    
    if any(word in filename_lower for word in ["evento", "event", "culto", "servicio"]):
        return "eventos"
    elif any(word in filename_lower for word in ["medicamento", "medicina", "farmacia", "medic"]):
        return "medicamentos"
    elif any(word in filename_lower for word in ["directorio", "lider", "pastor", "diacono"]):
        return "lideres"
    elif any(word in filename_lower for word in ["donacion", "donation", "ofrenda"]):
        return "donaciones"
    elif any(word in filename_lower for word in ["servicio", "service", "ministerio"]):
        return "servicios"
    else:
        return "documentos"

def get_title_from_filename(filename: str) -> str:
    """Obtener título del nombre del archivo"""
    import os
    name = os.path.splitext(filename)[0]
    return name.replace("_", " ").replace("-", " ").title()

def get_document_type(content_type: str) -> str:
    """Obtener tipo de documento basado en content type"""
    if content_type.startswith("text/"):
        return "texto"
    elif content_type == "application/pdf":
        return "pdf"
    elif "word" in content_type:
        return "word"
    elif "excel" in content_type or "spreadsheet" in content_type:
        return "excel"
    else:
        return "general"

def create_index_from_blob_storage(index_name: str, project_client, search_connection, index_client):
    """Crear índice desde Azure Blob Storage"""
    
    # Eliminar índice existente si existe
    try:
        existing_index = index_client.get_index(index_name)
        index_client.delete_index(index_name)
        logger.info(f"🗑️  Índice existente '{index_name}' eliminado")
    except Exception:
        logger.info(f"ℹ️  No se encontró índice existente '{index_name}'")
    
    # Crear definición del índice
    model = os.environ.get("EMBEDDINGS_MODEL", "text-embedding-ada-002")
    index_definition = create_index_definition(index_name, model)
    index_client.create_index(index_definition)
    logger.info(f"✅ Índice '{index_name}' creado")
    
    # Crear cliente de embeddings
    embeddings_client = project_client.inference.get_embeddings_client()
    
    # Crear documentos desde Blob Storage
    docs = create_docs_from_blob_storage(embeddings_client, model)
    
    if not docs:
        logger.warning("No se encontraron documentos en Blob Storage")
        return
    
    # Crear cliente de búsqueda
    search_client = SearchClient(
        endpoint=search_connection.endpoint_url,
        index_name=index_name,
        credential=AzureKeyCredential(key=search_connection.key),
    )
    
    # Subir documentos al índice
    search_client.upload_documents(docs)
    logger.info(f"➕ {len(docs)} documentos subidos al índice '{index_name}'")

def main():
    """Función principal"""
    print("🚀 VEA Connect - Crear Índice de Búsqueda con Azure AI Studio")
    print("=" * 60)
    
    try:
        # Verificar variables de entorno
        required_vars = ["AIPROJECT_CONNECTION_STRING", "AISEARCH_INDEX_NAME"]
        missing_vars = [var for var in required_vars if not os.getenv(var)]
        
        if missing_vars:
            print(f"❌ Variables de entorno faltantes: {', '.join(missing_vars)}")
            print("\n💡 Configura estas variables en tu archivo .env:")
            print("   AIPROJECT_CONNECTION_STRING=tu_connection_string")
            print("   AISEARCH_INDEX_NAME=vea-connect-index")
            return
        
        # Crear cliente de proyecto
        project_client = AIProjectClient.from_connection_string(
            conn_str=os.environ["AIPROJECT_CONNECTION_STRING"], 
            credential=DefaultAzureCredential()
        )
        logger.info("✅ Cliente de proyecto creado")
        
        # Obtener conexión de búsqueda
        search_connection = project_client.connections.get_default(
            connection_type=ConnectionType.AZURE_AI_SEARCH, 
            include_credentials=True
        )
        logger.info("✅ Conexión de búsqueda obtenida")
        
        # Crear cliente de índice
        index_client = SearchIndexClient(
            endpoint=search_connection.endpoint_url, 
            credential=AzureKeyCredential(key=search_connection.key)
        )
        
        # Crear índice
        index_name = os.environ["AISEARCH_INDEX_NAME"]
        
        create_index_from_blob_storage(index_name, project_client, search_connection, index_client)
        
        print("\n" + "=" * 60)
        print("🎉 ¡Índice creado exitosamente!")
        print(f"📋 Nombre del índice: {index_name}")
        print(f"🔍 Endpoint: {search_connection.endpoint_url}")
        print(f"📁 Contenedor Blob: {os.environ.get('STORAGE_CONTAINER_NAME', 'admin-documentos')}")
        print("\n📋 Próximos pasos:")
        print("1. Subir documentos al contenedor de Azure Blob Storage")
        print("2. Ejecutar este script nuevamente para reindexar")
        print("3. Probar búsquedas en el portal de Azure")
        print("4. Integrar con el chatbot")
        
    except Exception as e:
        logger.error(f"❌ Error general: {str(e)}")
        print(f"\n💡 Verifica tu configuración y vuelve a intentar")

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--index-name",
        type=str,
        help="Nombre del índice de búsqueda",
        default=os.environ.get("AISEARCH_INDEX_NAME", "vea-connect-index")
    )
    args = parser.parse_args()
    
    # Actualizar variables de entorno con argumentos
    os.environ["AISEARCH_INDEX_NAME"] = args.index_name
    
    main()
